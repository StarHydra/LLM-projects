import streamlit as st
import pdfplumber
import openpyxl
import json
import datetime
from groq import Groq
import pandas as pd
import os
from io import BytesIO


st.set_page_config(
    page_title="AI Document Structuring",
    page_icon="📄",
    layout="wide"
)

#Initializing Groq client
@st.cache_resource
def get_groq_client():
    api_key = os.getenv('GROQ_API_KEY') or st.secrets.get("GROQ_API_KEY")
    if not api_key:
        st.error("GROQ_API_KEY not found. Please configure it in Streamlit secrets.")
        st.stop()
    return Groq(api_key=api_key)

client = get_groq_client()

def extract_text(pdf_file):
    full_text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"
    return full_text.strip()

def extract_structured_data(text):
    prompt = """
You are an expert data extraction AI specializing in converting unstructured biographical narratives into structured key-value pairs.

Given the following text:

{text}

Your task:
- Dynamically identify all factual elements and group them into logical key-value pairs. Keys should be concise and descriptive (e.g., "First Name", "Date of Birth", "Current Salary", "Certifications 1").
- For values: Use exact original data where possible.
  - Dates: Output in YYYY-MM-DD format if mentioned (or infer from context like "June 15, 2002" -> "15-Jun-02").
  - Salaries: Numeric value without commas or currency (e.g., "350,000 INR" -> 350000 for salary, separate "INR" as "Salary Currency").
  - Add Company/Organization name as value where there is salary mentioned, add previous, current prefix to the key value depending upon date of joining.
  - If no company name is mentioned (or just terms like "first company","last company",etc. is used), then keep the value section blank
  - Add date of joining and of leaving as and where is mentioned
  - Percentages/Scores: Keep as it is (e.g., "92.5%" -> 92.5%, "8.7 on a n-point scale" -> 8.7 and then add the scale in comment section).
  - Keep units in key or value if integral (e.g., "35 years" for age).
  - For lists like certifications or skills, create sequential keys (e.g., "Certifications 1", "Certifications 2").
  - For certifications or skills, add the certification exam/company in the value.
  - For certifications or skills, add the year of certification and marks in the comment section.
- For comments: Pull relevant contextual sentences or phrases from the original text using exact wording. Include all descriptive details, explanations, or additional info here. If a section is purely descriptive (e.g., technical skills paragraph), use an empty value and put the full description in comments.
- Ensure 100% capture: No summarization, omission, or paraphrasing unless absolutely needed for a clean key-value (e.g., inferring "Birth City" from "born in Jaipur"). Preserve original sentence structure in comments.
- Do not introduce new information.
- Output ONLY a valid JSON array of objects in this exact format: [{{"key": "string", "value": "string or number as string", "comments": "string"}}]
- Order logically: personal info, professional, education, certifications, skills.

Make the JSON parsable and complete.
"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt.format(text=text)}],
        temperature=0.1,
        max_tokens=3000
    )

    try:
        json_str = response.choices[0].message.content.strip()
        if json_str.startswith('```json'):
            json_str = json_str[7:-3]
        elif json_str.startswith('```'):
            json_str = json_str[3:-3]
        data = json.loads(json_str)
        if not isinstance(data, list):
            raise ValueError("Not a list")
        return data
    except json.JSONDecodeError as e:
        raise ValueError(f"Failed to parse JSON from Groq response: {e}")

def parse_date(value):
    if not isinstance(value, str):
        return value
    value = value.strip()

    fmts = ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d-%B-%Y"]

    for fmt in fmts:
        try:
            return datetime.datetime.strptime(value, fmt).date()
        except ValueError:
            pass

    return value

def create_excel_output(data):
    #Creating Excel file in memory and returning as BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output"

    ws.append(["Sr No.", "Key", "Value", "Comments"])

    for row_idx, item in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=item.get("key", ""))

        raw_value = item.get("value", "")
        excel_value = parse_date(raw_value)
        cell = ws.cell(row=row_idx, column=3, value=excel_value)
        if isinstance(excel_value, datetime.date):
            cell.number_format = "DD-MMM-YY"

        ws.cell(row=row_idx, column=4, value=item.get("comments", ""))

    #Saving to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def clean_comments(data):
    df = pd.DataFrame(data)
    
    if 'comments' not in df.columns:
        return data
    
    comments = df['comments'].fillna('').astype(str).str.strip()
    
    df['comments'] = df['comments'].where(~df['comments'].duplicated(keep='last'), '')
    
    comments = df['comments'].copy()
    for i in range(len(df)):
        if not comments[i]:
            continue
        current = comments[i].lower()
        for j in range(len(df)):
            if i != j and comments[j] and current in comments[j].lower():
                if len(comments[i]) <= len(comments[j]):
                    df.at[i, 'comments'] = ''
                    break
    
    return df.to_dict('records')

#Streamlit UI
st.title("AI-Powered Document Structuring")
st.markdown("Transform unstructured PDF documents into structured Excel output using AI")

st.divider()

#Uploade pdf file
uploaded_file = st.file_uploader(
    "Upload PDF Document",
    type=['pdf'],
    help="Upload the PDF file you want to extract data from"
)

if uploaded_file:
    st.success(f"File uploaded: {uploaded_file.name}")
    
    #Button for processing
    if st.button("Extract & Structure Data", type="primary", use_container_width=True):
        try:
            with st.spinner("Processing your document"):
                progress = st.progress(0, "Extracting text from PDF")
                text = extract_text(uploaded_file)
                progress.progress(33, "Text extraction complete")
                
                progress.progress(33, "Analyzing document with AI")
                structured_data = extract_structured_data(text)
                progress.progress(66, f"Extracted {len(structured_data)} key-value pairs!")
                
                # Step 3: Clean comments
                progress.progress(66, "Cleaning and organizing data")
                cleaned_data = clean_comments(structured_data)
                
                # Step 4: Create Excel
                progress.progress(90, "Generating Excel file")
                excel_file = create_excel_output(cleaned_data)
                progress.progress(100, "Processing complete")
                
                st.success(f"Successfully extracted {len(cleaned_data)} data points")
                
                #Preview of data
                st.subheader("Data Preview")
                preview_df = pd.DataFrame(cleaned_data)
                st.dataframe(preview_df, use_container_width=True, height=400)
                
                #Button for download
                st.download_button(
                    label="Download Excel File",
                    data=excel_file,
                    file_name="Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
        except Exception as e:
            st.error(f"Error processing document: {str(e)}")
            st.exception(e)

else:
    st.info("Please upload a PDF file to get started")

#Sidebar with information
with st.sidebar:
    st.header("About")
    st.markdown("""
    This application uses AI to:
    - Extract text from PDF documents
    - Identify key-value pairs
    - Structure data into Excel format
    - Preserve original context and comments
    """)
    
    st.divider()
    
    st.header("Technical Stack")
    st.markdown("""
    - **PDF Processing:** pdfplumber
    - **AI Model:** Groq (Llama 3.3 70B)
    - **Excel Generation:** openpyxl
    - **Framework:** Streamlit
    """)
